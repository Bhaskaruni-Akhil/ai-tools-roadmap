#!/usr/bin/env python3
"""
Converts data/roadmap.xlsx → public/data.json

Usage:
  python scripts/excel_to_json.py
  python scripts/excel_to_json.py --input data/roadmap.xlsx --output public/data.json
"""

import argparse
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print('ERROR: openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
    sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_INPUT  = os.path.join(ROOT, 'data', 'roadmap.xlsx')
DEFAULT_OUTPUT = os.path.join(ROOT, 'public', 'data.json')


def cell_val(cell):
    """Return stripped string or None for empty cells."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip()
    return s if s else None


def parse_bool(v):
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().upper() in ('TRUE', '1', 'YES')


def parse_date(v):
    """Return ISO date string YYYY-MM-DD or None."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # Accept YYYY-MM-DD directly
    if len(s) == 10 and s[4] == '-':
        return s
    # openpyxl may return datetime objects
    try:
        from datetime import datetime, date
        if isinstance(v, (datetime, date)):
            return v.strftime('%Y-%m-%d')
    except Exception:
        pass
    return s or None


def parse_owners(v):
    """Return list of owner name strings."""
    if v is None:
        return []
    return [name.strip() for name in str(v).split(',') if name.strip()]


def sheet_to_rows(ws):
    """Return list of dicts keyed by header row."""
    rows = list(ws.iter_rows())
    if not rows:
        return []
    headers = [cell_val(c) for c in rows[0]]
    result = []
    for row in rows[1:]:
        values = [cell_val(c) for c in row]
        if all(v is None for v in values):
            continue  # skip blank rows
        result.append(dict(zip(headers, values)))
    return result


def validate_tools(tools_rows):
    required = {'id', 'name', 'category', 'status', 'impact'}
    errors = []
    for i, row in enumerate(tools_rows, start=2):
        missing = required - {k for k, v in row.items() if v is not None}
        if missing:
            errors.append(f'  Tools row {i}: missing required columns {missing}')
    if errors:
        print('Validation errors:\n' + '\n'.join(errors), file=sys.stderr)
        sys.exit(1)


def convert(input_path, output_path):
    print(f'Reading:  {input_path}')
    wb = load_workbook(input_path, data_only=True)

    if 'Tools' not in wb.sheetnames:
        print("ERROR: Sheet 'Tools' not found in workbook.", file=sys.stderr)
        sys.exit(1)
    if 'Milestones' not in wb.sheetnames:
        print("ERROR: Sheet 'Milestones' not found in workbook.", file=sys.stderr)
        sys.exit(1)
    if 'Subtasks' not in wb.sheetnames:
        print("ERROR: Sheet 'Subtasks' not found in workbook.", file=sys.stderr)
        sys.exit(1)

    tools_rows      = sheet_to_rows(wb['Tools'])
    milestones_rows = sheet_to_rows(wb['Milestones'])
    subtasks_rows   = sheet_to_rows(wb['Subtasks'])

    validate_tools(tools_rows)

    # Group milestones and subtasks by tool_id
    milestones_by_tool = {}
    for m in milestones_rows:
        tid = str(m.get('tool_id', '')).strip()
        milestones_by_tool.setdefault(tid, []).append(m)

    subtasks_by_tool = {}
    for s in subtasks_rows:
        tid = str(s.get('tool_id', '')).strip()
        subtasks_by_tool.setdefault(tid, []).append(s)

    tools = []
    for row in tools_rows:
        tid = str(row['id']).strip() if row['id'] is not None else ''

        milestones = [
            {
                'id':         str(m.get('id', '')).strip(),
                'label':      str(m.get('label', '')).strip(),
                'week':       int(float(str(m.get('week', 1)))) if m.get('week') is not None else 1,
                'done':       parse_bool(m.get('done')),
                'targetDate': parse_date(m.get('targetDate')),
            }
            for m in milestones_by_tool.get(tid, [])
        ]

        subtasks = [
            {
                'id':    str(s.get('id', '')).strip(),
                'label': str(s.get('label', '')).strip(),
                'done':  parse_bool(s.get('done')),
            }
            for s in subtasks_by_tool.get(tid, [])
        ]

        flagged = row.get('flaggedProgressWeek')
        flagged_val = int(float(str(flagged))) if flagged is not None else None

        tools.append({
            'id':                  tid,
            'name':                str(row.get('name', '')).strip(),
            'category':            str(row.get('category', '')).strip(),
            'status':              str(row.get('status', '')).strip(),
            'impact':              str(row.get('impact', '')).strip(),
            'description':         str(row.get('description', '') or '').strip(),
            'notes':               str(row.get('notes', '') or '').strip(),
            'w1Date':              parse_date(row.get('w1Date')),
            'targetEndDate':       parse_date(row.get('targetEndDate')),
            'flaggedProgressWeek': flagged_val,
            'owners':              parse_owners(row.get('owners')),
            'milestones':          milestones,
            'subtasks':            subtasks,
        })

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(tools, f, indent=2, ensure_ascii=False)

    print(f'Written:  {output_path}')
    print(f'Tools:    {len(tools)}')
    total_m = sum(len(t["milestones"]) for t in tools)
    total_s = sum(len(t["subtasks"])   for t in tools)
    print(f'Milestones: {total_m}  |  Subtasks: {total_s}')


def main():
    parser = argparse.ArgumentParser(description='Convert roadmap.xlsx to data.json')
    parser.add_argument('--input',  default=DEFAULT_INPUT,  help='Path to roadmap.xlsx')
    parser.add_argument('--output', default=DEFAULT_OUTPUT, help='Path to data.json')
    args = parser.parse_args()
    convert(args.input, args.output)


if __name__ == '__main__':
    main()
