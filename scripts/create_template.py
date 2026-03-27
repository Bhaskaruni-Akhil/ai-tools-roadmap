#!/usr/bin/env python3
"""
Run once to generate data/roadmap.xlsx with the 3 pre-loaded tools.
Usage: python scripts/create_template.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'roadmap.xlsx')

HEADER_FILL   = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT   = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL      = PatternFill('solid', fgColor='EBF3FB')
BORDER_SIDE   = Side(style='thin', color='CCCCCC')
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)


def style_header(ws, row, col_widths):
    for col_idx, width in enumerate(col_widths, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28


def style_data_row(ws, row, num_cols, alt=False):
    fill = ALT_FILL if alt else PatternFill('solid', fgColor='FFFFFF')
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill      = fill
        cell.border    = CELL_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 18


# ── Sheet 1: Tools ────────────────────────────────────────────────────────────
TOOLS_HEADERS = [
    'id', 'name', 'category', 'status', 'impact',
    'description', 'notes', 'w1Date', 'targetEndDate',
    'flaggedProgressWeek', 'owners',
]
TOOLS_WIDTHS = [6, 28, 14, 14, 10, 50, 30, 14, 14, 22, 24]

TOOLS_DATA = [
    [
        '1', 'HSD Conversion', 'hsd', 'In Progress', 'High',
        'Automated conversion of HSD artifacts to target format using AI',
        '', '2025-01-06', '2025-04-14', '', 'Akhil',
    ],
    [
        '2', 'Selenium Test Generation', 'selenium', 'In Progress', 'High',
        'AI-assisted Selenium test case generation via Windsurf + BT1 MCP server',
        '', '2025-01-06', '2025-05-05', '', 'Akhil',
    ],
    [
        '3', 'Playwright Conversion', 'playwright', 'Planning', 'Medium',
        'Convert Selenium/manual test suites to Playwright via AI-assisted migration',
        '', '2025-02-03', '2025-05-26', '', 'Akhil',
    ],
]

# ── Sheet 2: Milestones ───────────────────────────────────────────────────────
MILESTONES_HEADERS = ['tool_id', 'id', 'label', 'week', 'done', 'targetDate']
MILESTONES_WIDTHS  = [8, 8, 40, 8, 8, 14]

MILESTONES_DATA = [
    ['1', 'm1',  'Discovery & scoping',      1,  True,  ''],
    ['1', 'm2',  'Prototype ready',           4,  True,  ''],
    ['1', 'm3',  'Internal review',           7,  False, ''],
    ['1', 'm4',  'Team rollout',              10, False, ''],
    ['2', 'm5',  'MCP server integration',   1,  True,  ''],
    ['2', 'm6',  'XPath accuracy fix',       3,  True,  ''],
    ['2', 'm7',  'Demo to stakeholders',     5,  True,  ''],
    ['2', 'm8',  'XPath automation',         8,  False, ''],
    ['2', 'm9',  'Git integration',          11, False, ''],
    ['3', 'm10', 'Scope & feasibility',      2,  False, ''],
    ['3', 'm11', 'Conversion prototype',     6,  False, ''],
    ['3', 'm12', 'Pilot on 1 test suite',    9,  False, ''],
    ['3', 'm13', 'Full rollout',             13, False, ''],
]

# ── Sheet 3: Subtasks ─────────────────────────────────────────────────────────
SUBTASKS_HEADERS = ['tool_id', 'id', 'label', 'done']
SUBTASKS_WIDTHS  = [8, 8, 50, 8]

SUBTASKS_DATA = [
    ['1', 's1',  'Define input/output schema',              True ],
    ['1', 's2',  'Build AI parsing layer',                  True ],
    ['1', 's3',  'Error handling & edge cases',             False],
    ['1', 's4',  'QA & validation pass',                    False],
    ['2', 's5',  'Story/defect input parsing',              True ],
    ['2', 's6',  'Reliable XPath capture via Selenium IDE', True ],
    ['2', 's7',  'Automate XPath recording',                False],
    ['2', 's8',  'Git integration for test export',         False],
    ['3', 's9',  'Audit existing Selenium suite',           False],
    ['3', 's10', 'Define Playwright target format',         False],
    ['3', 's11', 'Build AI conversion script',              False],
    ['3', 's12', 'Validate converted tests',                False],
]


def write_sheet(ws, headers, widths, data):
    ws.freeze_panes = 'A2'
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = h
    style_header(ws, 1, widths)

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = val
        style_data_row(ws, row_idx, len(headers), alt=(row_idx % 2 == 0))


def main():
    wb = Workbook()

    ws_tools = wb.active
    ws_tools.title = 'Tools'
    write_sheet(ws_tools, TOOLS_HEADERS, TOOLS_WIDTHS, TOOLS_DATA)

    ws_milestones = wb.create_sheet('Milestones')
    write_sheet(ws_milestones, MILESTONES_HEADERS, MILESTONES_WIDTHS, MILESTONES_DATA)

    ws_subtasks = wb.create_sheet('Subtasks')
    write_sheet(ws_subtasks, SUBTASKS_HEADERS, SUBTASKS_WIDTHS, SUBTASKS_DATA)

    out = os.path.abspath(OUTPUT_PATH)
    wb.save(out)
    print(f'Created: {out}')
    print()
    print('Columns reference:')
    print('  Tools      — id | name | category | status | impact | description | notes | w1Date | targetEndDate | flaggedProgressWeek | owners')
    print('  Milestones — tool_id | id | label | week | done | targetDate')
    print('  Subtasks   — tool_id | id | label | done')
    print()
    print('  category values : hsd | selenium | playwright')
    print('  status values   : In Progress | Planning | Backlog | Done')
    print('  impact values   : High | Medium | Low')
    print('  done values     : TRUE | FALSE')
    print('  date format     : YYYY-MM-DD  (leave blank for none)')
    print('  owners          : comma-separated names  e.g. "Akhil, Jacob"')


if __name__ == '__main__':
    main()
